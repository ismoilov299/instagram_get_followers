import sqlite3
import json
import asyncio
import random
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from aiogram import Router, F
from aiogram.types import Message, CallbackQuery, FSInputFile
from aiogram.filters import Command, StateFilter
from aiogram.fsm.context import FSMContext

from bot.states import InstagramStates
from bot.keyboards import get_followers_keyboard, get_winner_keyboard, get_export_keyboard
from services.instagram_api import InstagramAPI

router = Router()

# Путь к файлу базы данных SQLite
DATABASE_PATH = "instagram_followers.db"

# Фиксированный Instagram username
FIXED_INSTAGRAM_USERNAME = "zayd.catlover"


# Функции для работы с базой данных SQLite
async def initialize_database():
    """Инициализация базы данных SQLite"""
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()

    # Создаем таблицу для хранения информации об аккаунтах
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS accounts (
        username TEXT PRIMARY KEY,
        followers_count INTEGER,
        full_name TEXT,
        following_count INTEGER,
        posts_count INTEGER,
        bio TEXT,
        update_timestamp INTEGER
    )
    ''')

    # Создаем таблицу для хранения подписчиков
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS followers (
        id TEXT,
        username TEXT,
        link TEXT,
        account_username TEXT,
        PRIMARY KEY (id, account_username),
        FOREIGN KEY (account_username) REFERENCES accounts (username)
    )
    ''')

    conn.commit()
    conn.close()


async def save_followers_to_db(user_info, followers_list):
    """Сохранить данные о подписчиках в базе данных SQLite"""
    username = user_info['username']
    followers_count = user_info['followers_count']

    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()

    # Начинаем транзакцию
    conn.execute("BEGIN TRANSACTION")

    try:
        # Удаляем старые данные об аккаунте, если они есть
        cursor.execute("DELETE FROM accounts WHERE username = ?", (username,))

        # Вставляем новую информацию об аккаунте
        cursor.execute('''
        INSERT INTO accounts (username, followers_count, full_name, following_count, posts_count, bio, update_timestamp)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (
            username,
            followers_count,
            user_info.get('full_name', ''),
            user_info.get('following_count', 0),
            user_info.get('posts_count', 0),
            user_info.get('bio', ''),
            int(asyncio.get_event_loop().time())
        ))

        # Удаляем старых подписчиков этого аккаунта
        cursor.execute("DELETE FROM followers WHERE account_username = ?", (username,))

        # Вставляем новых подписчиков
        for follower in followers_list:
            cursor.execute('''
            INSERT INTO followers (id, username, link, account_username)
            VALUES (?, ?, ?, ?)
            ''', (
                follower['id'],
                follower['username'],
                follower['link'],
                username
            ))

        # Завершаем транзакцию
        conn.commit()
        return True

    except Exception as e:
        # Откатываем транзакцию в случае ошибки
        conn.rollback()
        print(f"Ошибка при сохранении данных в базу: {e}")
        return False

    finally:
        conn.close()


async def get_account_info_from_db(username):
    """Получить информацию об аккаунте из базы данных"""
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()

    cursor.execute('''
    SELECT username, followers_count, full_name, following_count, posts_count, bio, update_timestamp
    FROM accounts
    WHERE username = ?
    ''', (username,))

    result = cursor.fetchone()
    conn.close()

    if not result:
        return None

    return {
        'username': result[0],
        'followers_count': result[1],
        'full_name': result[2],
        'following_count': result[3],
        'posts_count': result[4],
        'bio': result[5],
        'update_timestamp': result[6]
    }


async def get_followers_from_db(username):
    """Получить список подписчиков аккаунта из базы данных"""
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()

    cursor.execute('''
    SELECT id, username, link
    FROM followers
    WHERE account_username = ?
    ''', (username,))

    followers = []
    for row in cursor.fetchall():
        followers.append({
            'id': row[0],
            'username': row[1],
            'link': row[2]
        })

    conn.close()
    return followers


@router.message(Command("start"))
async def cmd_start(message: Message, state: FSMContext, instagram_api: InstagramAPI):
    """
    Начало работы бота. Теперь сразу используется фиксированный пользователь.
    """
    # Инициализируем базу данных при первом запуске
    await initialize_database()

    # Показываем приветственное сообщение
    await message.answer(
        f"👋 Assalomu alaykum! Instagram follower bot'ga xush kelibsiz!\n\n"
        f"🔍 Bot faqat @{FIXED_INSTAGRAM_USERNAME} profilidan ma'lumot oladi."
    )

    # Передаем instagram_api в функцию
    await process_fixed_user(message, state, instagram_api)


@router.message(Command("followers"))
async def cmd_followers(message: Message, state: FSMContext, instagram_api: InstagramAPI):
    """
    Команда для повторного получения подписчиков
    """
    # Передаем instagram_api в функцию
    await process_fixed_user(message, state, instagram_api)


async def process_fixed_user(message: Message, state: FSMContext, instagram_api: InstagramAPI):
    """
    API limit bo'lsa avval bazadan ma'lumot olish
    """
    username = FIXED_INSTAGRAM_USERNAME

    # Показываем, что бот начал работу
    await message.answer(f"🔍 @{username} profili tekshirilmoqda...")
    await message.bot.send_chat_action(chat_id=message.chat.id, action="typing")

    # Avval bazadan ma'lumot olishga harakat qilamiz
    db_user_info = await get_account_info_from_db(username)
    db_followers = await get_followers_from_db(username)

    user_info = None
    api_working = False

    # API ni sinab ko'rish
    try:
        user_info = await instagram_api.get_user_info(username)
        if user_info:
            api_working = True
            print("API is working, got fresh user info")
    except Exception as e:
        error_message = str(e).lower()
        if "429" in error_message or "quota" in error_message or "limit" in error_message:
            print("API quota exceeded, using database data")
        else:
            print(f"API Error: {e}")

    # Agar API ishlamasa va bazada ma'lumot bo'lsa, bazadan foydalanish
    if not api_working and db_user_info:
        print(f"Using database user info for {username}")

        # Bazadagi ma'lumotlarni API formatiga o'tkazish
        user_info = {
            'id': str(db_user_info.get('username', username)),  # ID o'rniga username ishlatamiz
            'username': db_user_info['username'],
            'full_name': db_user_info['full_name'],
            'followers_count': db_user_info['followers_count'],
            'following_count': db_user_info['following_count'],
            'posts_count': db_user_info['posts_count'],
            'bio': db_user_info['bio']
        }

        # Ma'lumotni ko'rsatish (API limit haqida aytmaslik)
        await message.answer(
            f"✅ Ma'lumotlar topildi!\n\n"
            f"👤 *{user_info['full_name']}* (@{user_info['username']})\n"
            f"📊 Statistika:\n"
            f"- Obunachilar: {user_info['followers_count']}\n"
            f"- Obuna bo'lganlar: {user_info['following_count']}\n"
            f"- Postlar: {user_info['posts_count']}\n\n"
            f"🔗 Link: https://www.instagram.com/{user_info['username']}\n\n"
            f"Obunachilarni yuklash boshlanmoqda...",
            parse_mode="Markdown"
        )

        # Agar bazada followers ham bo'lsa
        if db_followers:
            print(f"Found {len(db_followers)} followers in database")

            # State ga ma'lumotlarni saqlash
            await state.update_data(
                instagram_user=user_info,
                followers_list=db_followers,
                total_fetched=len(db_followers),
                total_followers=user_info['followers_count'],
                is_database_data=True
            )

            # Haqiqiydek yuklash simulyatsiyasi
            status_message = await message.answer("🔄 Obunachilar yuklanmoqda... 0/0")
            await simulate_database_loading_realistic(
                message, status_message.message_id, len(db_followers), user_info['followers_count']
            )

            # G'olib tanlash tugmasi
            await message.answer(
                "G'olibni aniqlash uchun tugmani bosing:",
                reply_markup=get_winner_keyboard()
            )
            return
        else:
            # Bazada followers yo'q
            await state.update_data(instagram_user=user_info)
            await message.answer(
                "⚠️ Saqlangan obunachilar ma'lumoti topilmadi.\n"
                "API limit tugaganidan keyin qayta urinib ko'ring."
            )
            return

    # Agar API ishlamasa va bazada ham ma'lumot bo'lmasa
    elif not api_working and not db_user_info:
        await message.answer(
            f"❌ API limitlari tugagan va '@{username}' uchun saqlangan ma'lumot topilmadi.\n"
            f"Iltimos, keyinroq qayta urinib ko'ring."
        )
        return

    # Agar API ishlasa, odatiy yo'l bilan davom etish
    elif api_working and user_info:
        print("API is working, proceeding with normal flow")

        # Сохраняем информацию о пользователе
        await state.update_data(instagram_user=user_info)

        # Проверяем, нужно ли обновлять данные
        need_update = False

        if not db_user_info:
            need_update = True
        else:
            followers_diff = abs(db_user_info['followers_count'] - user_info['followers_count'])
            if followers_diff >= 1000:
                need_update = True

        # Если обновление не требуется, используем кэшированные данные
        if not need_update and db_user_info and db_followers:
            await message.answer(
                f"👤 *{user_info['full_name']}* (@{user_info['username']})\n"
                f"📊 Statistika:\n"
                f"- Obunachilar: {user_info['followers_count']}\n"
                f"- Obuna bo'lganlar: {user_info['following_count']}\n"
                f"- Postlar: {user_info['posts_count']}\n\n"
                f"🔗 Link: https://www.instagram.com/{user_info['username']}\n\n",
                parse_mode="Markdown"
            )

            await state.update_data(
                followers_list=db_followers,
                total_fetched=len(db_followers),
                total_followers=db_user_info['followers_count']
            )

            await message.answer(
                "G'olibni aniqlash uchun tugmani bosing:",
                reply_markup=get_winner_keyboard()
            )
        else:
            # Yangi ma'lumot yuklash kerak
            if db_user_info:
                followers_diff = abs(db_user_info['followers_count'] - user_info['followers_count'])
                await message.answer(
                    f"🔄 Obunachilar soni {followers_diff} ta o'zgargan, yangi ma'lumotlar yuklanmoqda...",
                    parse_mode="Markdown"
                )

            await message.answer(
                f"✅ Ma'lumotlar topildi!\n\n"
                f"👤 *{user_info['full_name']}* (@{user_info['username']})\n"
                f"📊 Statistika:\n"
                f"- Obunachilar: {user_info['followers_count']}\n"
                f"- Obuna bo'lganlar: {user_info['following_count']}\n"
                f"- Postlar: {user_info['posts_count']}\n\n"
                f"🔗 Link: https://www.instagram.com/{user_info['username']}\n\n"
                f"Obunachilarni yuklash boshlanmoqda...",
                parse_mode="Markdown"
            )

            status_message = await message.answer("🔄 Obunachilar yuklanmoqda... 0/0")

            await state.update_data(
                current_user_id=user_info['id'],
                followers_list=[],
                next_max_id=None,
                total_fetched=0,
                total_followers=user_info['followers_count'],
                status_message_id=status_message.message_id
            )

            await fetch_all_followers(message, state, instagram_api)


async def simulate_database_loading_realistic(message, status_message_id: int, actual_count: int,
                                              target_followers: int):
    """
    Tezlashtirilgan bazadan yuklash simulyatsiyasi
    """
    # Katta batch lar ishlatish - tezroq yuklash uchun
    batch_size = 200  # 50 dan 200 ga ko'tarildi
    loaded = 0
    batch_count = 0

    # Kamroq qadam bilan yuklash
    while loaded < actual_count:
        batch_count += 1
        await asyncio.sleep(random.uniform(0.1, 0.3))  # 0.5-1.0 dan 0.1-0.3 ga kamaytirildi

        remaining = actual_count - loaded
        current_batch_size = min(batch_size, remaining)
        loaded += current_batch_size

        # target_followers ga nisbatan percentage ko'rsatish
        percentage = min(100, int((loaded / target_followers) * 100))

        # Har 3-batch da bir marta status yangilash (tezroq)
        if batch_count % 3 == 0 or loaded >= actual_count:
            try:
                await message.bot.edit_message_text(
                    text=f"🔄 Obunachilar yuklanmoqda... {loaded}/{target_followers} ({percentage}%)",
                    chat_id=message.chat.id,
                    message_id=status_message_id
                )
            except Exception:
                pass

        # Kamroq typing action
        if batch_count % 5 == 0:  # Har 5-batch da bir marta
            await message.bot.send_chat_action(chat_id=message.chat.id, action="typing")

        batch_size = random.randint(150, 250)  # Katta batch hajmi

    # Final status
    final_percentage = min(100, int((loaded / target_followers) * 100))
    try:
        await message.bot.edit_message_text(
            text=f"✅ Obunachilar yuklandi",
            chat_id=message.chat.id,
            message_id=status_message_id
        )
    except Exception:
        pass


@router.callback_query(F.data == "select_winner")
async def select_winner(callback: CallbackQuery, state: FSMContext):
    """
    G'olib tanlash - bazadagi ma'lumot bilan ham ishlaydi
    """
    await callback.answer("🎲 G'olib tanlanmoqda...")

    data = await state.get_data()
    followers_list = data.get('followers_list', [])
    total_fetched = data.get('total_fetched', 0)
    is_database_data = data.get('is_database_data', False)

    if not followers_list:
        await callback.message.answer("❌ G'olibni aniqlash uchun obunachilar ro'yxati mavjud emas!")
        return

    # Animatsiya
    await callback.bot.send_chat_action(chat_id=callback.message.chat.id, action="typing")
    await asyncio.sleep(1.5)

    winner = random.choice(followers_list)
    winner_index = followers_list.index(winner) + 1

    dots_message = await callback.message.answer("🎲 G'olib tanlanmoqda...")

    animation_texts = [
        "🎲 G'olib tanlanmoqda...",
        "🎲 G'olib hisoblanmoqda...",
        "🎲 Natijalar tayyorlanmoqda..."
    ]

    for text in animation_texts:
        await asyncio.sleep(0.7)
        try:
            await dots_message.edit_text(text)
        except Exception:
            continue

    await asyncio.sleep(0.7)

    winner_text = (
        f"🎉 *G'OLIB ANIQLANDI!* 🎉\n\n"
        f"🏆 G'olib: [{winner['username']}]({winner['link']})\n"
        f"🔢 G'olibning tartib raqami: {winner_index} \n\n"
        f"Tabriklaymiz! 🎊"
    )

    try:
        await dots_message.delete()
    except Exception:
        pass

    await callback.message.answer(
        winner_text,
        parse_mode="Markdown",
        disable_web_page_preview=True
    )

    # Boshqa g'olib tanlash
    await callback.message.answer(
        "Boshqa g'olibni aniqlash uchun tugmani bosing:",
        reply_markup=get_winner_keyboard()
    )


async def fetch_all_followers(message: Message, state: FSMContext, instagram_api: InstagramAPI):
    """
    Haqiqiy API bilan followers yuklash
    """
    data = await state.get_data()
    username = data.get('instagram_user', {}).get('username', '')
    status_message_id = data.get('status_message_id')
    total_followers = data.get('total_followers')
    user_info = data.get('instagram_user')

    followers_list = []
    pagination_token = None
    total_fetched = 0
    last_status_text = ""
    batch_count = 0

    # Функция безопасного обновления сообщения
    async def update_status_safely(text):
        nonlocal last_status_text
        if text == last_status_text:
            return
        last_status_text = text
        await safe_edit_message(message.bot, message.chat.id, status_message_id, text)

    # Начинаем загрузку
    while total_fetched < total_followers:
        batch_count += 1
        await message.bot.send_chat_action(chat_id=message.chat.id, action="typing")

        try:
            batch_result = await instagram_api.get_user_followers_batch(
                username, 50, pagination_token
            )

            if not batch_result or not batch_result.get('followers'):
                await update_status_safely("⚠️ Obunachilarni yuklashda xatolik yuz berdi.")
                break

            new_followers = batch_result.get('followers', [])
            followers_list.extend(new_followers)
            total_fetched = len(followers_list)
            pagination_token = batch_result.get('next_max_id')

            percentage = min(100, int((total_fetched / total_followers) * 100))
            await update_status_safely(
                f"🔄 Obunachilar yuklanmoqda... {total_fetched}/{total_followers} ({percentage}%) - Batch {batch_count}")

            if not pagination_token or not new_followers or not batch_result.get('has_more', True):
                await update_status_safely(f"✅ Barcha obunachilar yuklandi")
                break

            await asyncio.sleep(0.8)

            if batch_count > 2000:
                await update_status_safely(f"⚠️ Xavfsizlik chegarasiga yetdi: {total_fetched} ta obunachi yuklandi")
                break

        except Exception as e:
            print(f"Error fetching followers: {e}")
            await asyncio.sleep(3)
            continue

    # Сохраняем результаты
    await state.update_data(
        followers_list=followers_list,
        total_fetched=total_fetched,
        is_database_data=False
    )

    # Сохраняем в базу
    if user_info and followers_list:
        save_success = await save_followers_to_db(user_info, followers_list)
        if save_success:
            print(f"Successfully saved {len(followers_list)} followers to database")

    # Показываем итоговый статус
    final_percentage = min(100, int((total_fetched / total_followers) * 100))
    await update_status_safely(f"✅ Obunachilar yuklandi")

    # Предлагаем выбрать победителя
    if followers_list:
        await message.answer(
            "G'olibni aniqlash uchun tugmani bosing:",
            reply_markup=get_winner_keyboard()
        )
    else:
        await message.answer("❌ Obunachilar ro'yxatini olib bo'lmadi.")


@router.callback_query(F.data == "export_excel")
async def export_to_excel(callback: CallbackQuery, state: FSMContext):
    """
    Export followers list to Excel file
    """
    await callback.answer("📊 Excel fayl tayyorlanmoqda...")

    # Get the followers list from state
    data = await state.get_data()
    followers_list = data.get('followers_list', [])
    total_fetched = data.get('total_fetched', 0)

    if not followers_list:
        await callback.message.answer("❌ Eksport qilish uchun obunachilar ro'yxati mavjud emas!")
        return

    # Show typing action to indicate processing
    await callback.bot.send_chat_action(chat_id=callback.message.chat.id, action="upload_document")

    # Create Excel file in memory
    excel_file = await create_excel_file(followers_list, FIXED_INSTAGRAM_USERNAME)

    # Send the file to user
    await callback.message.answer_document(
        FSInputFile(
            excel_file,
            filename=f"{FIXED_INSTAGRAM_USERNAME}_followers.xlsx"
        ),
        caption=f"📊 {FIXED_INSTAGRAM_USERNAME} uchun {total_fetched} ta obunachi ma'lumotlari."
    )

    # Allow selecting a winner
    await callback.message.answer(
        "🎲 G'olibni aniqlash uchun tugmani bosing:",
        reply_markup=get_winner_keyboard()
    )


async def create_excel_file(followers_list, account_username):
    """
    Create Excel file with followers data
    """
    # Create a workbook and select active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Obunachilar"

    # Define header style
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    # Add headers
    headers = ["№", "Username", "Instagram Link", "ID"]
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Add data
    for row_num, follower in enumerate(followers_list, 2):
        worksheet.cell(row=row_num, column=1).value = row_num - 1  # № (counter)
        worksheet.cell(row=row_num, column=2).value = follower['username']
        worksheet.cell(row=row_num, column=3).value = follower['link']
        worksheet.cell(row=row_num, column=4).value = follower['id']

    # Auto-adjust column width
    for column_cells in worksheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = length + 5

    # Save to in-memory file
    file_path = f"temp_{account_username}_followers.xlsx"
    workbook.save(file_path)

    return file_path


async def safe_edit_message(bot, chat_id, message_id, text, **kwargs):
    """
    Безопасно обновляет сообщение, игнорируя ошибки "message is not modified"
    """
    try:
        await bot.edit_message_text(
            text=text,
            chat_id=chat_id,
            message_id=message_id,
            **kwargs
        )
        return True
    except Exception as e:
        # Игнорируем ошибку "message is not modified"
        if "message is not modified" in str(e).lower():
            return True  # Сообщение уже содержит нужный текст

        # Для других ошибок просто логируем и продолжаем
        print(f"Ошибка при обновлении сообщения: {e}")
        return False